#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Xiang Wang <ramwin@qq.com>


import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook


logger = logging.getLogger(__name__)


def convert(input_path, output_path, from_column, to_column):
    assert Path(input_path).exists(), f"文件: {input_path} 不存在"
    assert not Path(output_path).exists(), f"文件: {output_path} 已存在"
    wb = load_workbook(input_path)
    ws = wb.active
    for row in range(1, ws.max_row+1):
        if ws[f"{to_column}{row}"].value:
            logger.debug(f"{to_column}{row} 有东西")
            continue
        value = ws[f"{from_column}{row}"].value
        logger.debug(f"复制{value}到{to_column}{row}")
        ws[f"{to_column}{row}"].value = value
    wb.save(output_path)
