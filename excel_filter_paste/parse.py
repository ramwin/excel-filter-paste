#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Xiang Wang <ramwin@qq.com>


import datetime
import logging

import pandas
import re

from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def get_type(string):
    """
    返回括号内的数值
    Parameters:
        "2022-04 (aabb-cc) AY8A"
    Returns:
        "aabb-cc"
    """
    try:
        return re.split(r'[\(\)（）]', string)[1]
    except Exception:
        raise Exception(f"{string}格式不正确")


def get_sheet_type(path):
    """
    返回丰通的dict, 不包含隐藏的sheet
    Returns:
        {
            "sheet_name": "typename"
        }
    """
    wb = load_workbook(path)
    result = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        if sheet.sheet_state == 'hidden':
            continue
        _type = get_type(sheet["A3"].value)
        result[sheetname] = _type
    return result
