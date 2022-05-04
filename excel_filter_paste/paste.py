#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Xiang Wang <ramwin@qq.com>


import datetime
import logging
from pathlib import Path

from month import XMonth

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

import pandas

from .parse import get_sheet_type


logger = logging.getLogger(__name__)


class Convert:

    def __init__(self, input_path, output_path, directory):
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        self.directory = Path(directory)
        self.check()

    def check(self):
        assert self.directory.exists(), f"{self.directory}不存在"
        assert self.input_path.exists(), f"{self.input_path}不存在"

    @classmethod
    def paste_convert(cls, input_path, output, directory):
        cls(input_path, output, directory).run()

    def run(self):
        self.load_data()
        Guangqifengtian(
            sheet=self.sheet,
            directory=self.directory,
        ).run()
        Changchunfengyue(
            sheet=self.sheet,
            directory=self.directory,
        ).run()
        SichuanConvert(
            sheet=self.sheet,
            directory=self.directory,
        ).run()
        TianjinConvert(
            sheet=self.sheet,
            directory=self.directory,
        ).run()
        self.wb.save(self.output_path)

    def load_data(self):
        self.wb = openpyxl.load_workbook(self.input_path)
        self.sheet = self.wb.active


class BaseConvert:
    FILE_PATTERN = None
    SUFFIX = ".xlsx"

    def __init__(self, sheet, directory):
        self.sheet = sheet
        self.directory = directory

    def run(self):
        self.data = self.get_data()
        self.paste()

    def get_file(self):
        name = self.FILE_PATTERN
        directory = self.directory
        if not name:
            raise NotImplementedError
        pattern = f"*{name}*{self.SUFFIX}"
        logger.info(f"寻找文件: {pattern}")
        files = list(directory.glob(pattern))
        if len(files) == 1:
            return files[0]
        if len(files) == 0:
            return f"{directory}里面不存在 {name} 的xlsx文件"
        return f"{directory}里面有不止1个{name}的xlsx文件"

    def get_data(self):
        data = self.parse_data(self.get_file())
        return data

    def paste(self):
        ws = self.sheet
        data = self.data
        # 收获数
        for _type, _type_data in data.items():
            position = self.get_position(_type)
            logger.debug(f"粘贴数据{_type_data}")
            for date, value in _type_data.items():
                for key, number in value.items():
                    row = position[key]
                    column = get_column_letter(
                        column_index_from_string(position["column"]) + date.day - 1
                    )
                    if number and not pandas.isna(number):
                        logger.info(f"修改{column}{row}为{number}")
                        ws[f"{column}{row}"].value = number

    def parse_data(self, path):
        raise NotImplementedError


class SichuanConvert(BaseConvert):
    FILE_PATTERN = "库存滚动日报表_2022"

    def parse_data(self, path, end=None):
        """
        解析四川一汽丰通的日报
        Parameters:
            path: 文件路径
            end: datetime.date
        Returns:
            {
                {
                    "type1": {
                        Timestamp("2022-04-01"): {
                            "D": 1,
                            "收货数": 2,
                            "发货数": 3,
                            "其他出货 4,
                        }
                    }
                }
            }
        """
        sheet_name_type = get_sheet_type(path)
        if end is None:
            end = datetime.date.today()
        end = pandas.Timestamp(end)
        df_dict = pandas.read_excel(
            path, names=['日期', None, None, None, "收货数", "发货数MP", "其他出货"],
            header=9,
            index_col="日期",
            usecols=["日期", "收货数", "发货数MP", "其他出货"],
            sheet_name=list(sheet_name_type.keys()),
        )
        result = {}
        for sheet_name, df in df_dict.items():
            df = df[df.index.notnull()]
            df = df[~pandas.to_datetime(df.index, errors="coerce").isnull()]
            _type = sheet_name_type[sheet_name]
            result[_type] = df[df.index < end].to_dict(orient="index")
        return result

    def get_position(self, _type):
        """
        Parameters:
            _type: 部品番号 4265A-0N011
        Return:
            粘贴位置 {
                "收获数": 2248(row),
                "发货数": 2249(1 index row),
                "其他出货": 2250,
                "column": "I"
            }
        """
        ws = self.sheet
        result = {"column": "I"}
        e_index = column_index_from_string('E') - 1
        _type_index = column_index_from_string('D') - 1
        column_index = column_index_from_string('G') - 1
        for index, row in enumerate(ws.rows, 1):
            e_value = row[e_index].value
            if e_value \
                    and "四川成都" in e_value \
                    and "丰通" in e_value \
                    and row[_type_index].value == _type:
                if row[column_index].value in ["收货数", "发货数MP", "其他出货"]:
                    result[row[column_index].value] = index
        logger.info(result)
        return result


class TianjinConvert(BaseConvert):
    FILE_PATTERN = "在库表-YH 2022"

    def get_position(self, _type):
        """
        Parameters:
            _type: 部品番号 4265A-0N011
        Return:
            粘贴位置 {
                "收获数": 2248(row),
                "发货数": 2249(1 index row),
                "其他出货": 2250,
                "column": "I"
            }
        """
        ws = self.sheet
        result = {"column": "I"}
        e_index = column_index_from_string('E') - 1
        _type_index = column_index_from_string('D') - 1
        column_index = column_index_from_string('G') - 1
        for index, row in enumerate(ws.rows, 1):
            e_value = row[e_index].value
            if e_value \
                    and "一汽丰田" in e_value \
                    and "丰通" in e_value \
                    and row[_type_index].value == _type:
                if row[column_index].value in ["收货数", "发货数MP", "其他出货"]:
                    result[row[column_index].value] = index
        logger.info(result)
        return result

    def parse_data(self, path, end=None):
        """
        解析天津一汽丰通的日报
        Parameters:
            path: 文件路径
            end: datetime.date
        Returns:
            {
                {
                    "type1": {
                        Timestamp("2022-04-01"): {
                            "D": 1,
                            "收货数": 2,
                            "发货数": 3,
                            "其他出货 4,
                        }
                    }
                }
            }
        """
        sheet_name_type = get_sheet_type(path)
        if end is None:
            end = datetime.date.today()
        end = pandas.Timestamp(end)
        df_dict = pandas.read_excel(
            path,
            names=[
                '日期', None, None, "D", "收货数",
                None, None, None, None, None,
                "发货数MP", "其他出货"],
            header=9,
            index_col="日期",
            usecols=["日期", "收货数", "发货数MP", "其他出货"],
            sheet_name=list(sheet_name_type.keys()),
        )
        result = {}
        for sheet_name, df in df_dict.items():
            df = df[df.index.notnull()]
            df = df[~pandas.to_datetime(df.index, errors="coerce").isnull()]
            _type = sheet_name_type[sheet_name]
            result[_type] = df[df.index < end].to_dict(orient="index")
        return result


class Changchunfengyue(BaseConvert):
    FILE_PATTERN = "轮胎日报表"
    name = "长春丰越"

    def get_sheet_name(self, path):
        """
        Returns:
            year: 2022
            month: 4
            sheet_name: 202204
        """
        wb = load_workbook(path)
        wb.sheetnames.sort()
        year_month = sorted(wb.sheetnames)[-1]
        year = int(year_month[0:4])
        month = int(year_month[4:6])
        return year, month, year_month

    def parse_data(self, path, end=None):
        """
        解析丰越的日报
        Parameters:
            path: 文件路径
            end: datetime.date
        Returns:
            {
                {
                    "type1": {  # 这里的type1是直接的type
                        Timestamp("2022-04-01"): {
                            "D": 1,
                            "收货数": 2,
                            "发货数": 3,
                            "其他出货 4,
                        }
                    }
                }
            }
        """
        logger.info("解析丰越的日报")
        year, month, sheet_name = self.get_sheet_name(path)
        df = pandas.read_excel(
            path, header=[1, 2, 3],
            sheet_name=sheet_name,
        )
        df = df[
            ~pandas.to_numeric(df[df.columns[0]],
                               errors="coerce").isnull()
        ]
        df[df.columns[0]] = df[df.columns[0]].map(lambda x: int(x))
        df = df[
            df[df.columns[0]] <= XMonth(year, month).last_date().day
        ]
        assert df.columns[0][-1] == '日期', f"{path}的格式不对，可能是A2单元格没有像以前一样合并了"
        df["日期"] = df[df.columns[0]].map(
            lambda x: pandas.Timestamp(
                f"{year}-{month}-{x}"
            )
        )
        df.set_index("日期", inplace=True)
        df = df['丰越品']
        # {'H1829': '***H1829', 'R5250', 'R5251'}
        # _type_set = ...很长\nR5251
        _type_set = {i[0] for i in df.columns}
        result = {}
        for _type in _type_set:
            result.update(self.get_result(df, _type))
        logger.debug(result)
        return result

    def get_result(self, df, _type):
        real_type = _type.split("\n")[1]
        data = df[_type].filter(['出库', '入库']).rename(
            columns={
                "入库": "收货数",
                "出库": "发货数MP"
            },
        ).to_dict(orient='index')
        return {real_type: data}

    def get_position(self, _type):
        ws = self.sheet
        result = {"column": "I"}
        e_index = column_index_from_string('E') - 1
        _type_index = column_index_from_string('B') - 1
        column_index = column_index_from_string('G') - 1
        for index, row in enumerate(ws.rows, 1):
            e_value = row[e_index].value
            if e_value \
                    and "长春丰越" == e_value \
                    and row[_type_index].value == _type:
                if row[column_index].value in ["收货数", "发货数MP"]:
                    result[row[column_index].value] = index
        logger.info(result)
        return result


class Guangqifengtian(BaseConvert):
    FILE_PATTERN = "在库联络表"
    SUFFIX = ".xlsm"

    def parse_data(self, path, end=None):
        df = pandas.read_excel(
            path, header=[2, 3, 4, 5, 6],
        )
        date_column = df.columns[1]
        df = df[~pandas.to_datetime(df[date_column], errors="coerce").isnull()]
        df.set_index(date_column, inplace=True)
        _type_set = {
            i[0]
            for i in df.columns
            if isinstance(i[0], str) and 11 <= len(i[0]) and "Unnamed" not in i[0]
        }
        logger.info(f"广汽丰田有的型号: {_type_set}")
        result = {}
        for _type in _type_set:
            result.update(self.get_result(df, _type))
        logger.debug(result)
        return result

    def get_result(self, df, _type):
        real_type = _type[-6:-1]
        df = df[_type]
        # ["入库", "undefined"]
        new_df = pandas.DataFrame(index=df.index)
        for title, value in {
            "入库": "收货数",
            "实绩": "发货数MP",
            "返厂数": "不良品返厂",
        }.items():
            new_df[value] = self.get_series(
                df,
                self.get_columns(df, title)
            )
        data = new_df.to_dict(orient='index')
        return {
            real_type: data
        }

    def get_columns(self, df, title):
        """
        Parameters:
            title: "入库"
        Return:
            columns: [
                '入库', 'Unnamed: 2_level_2', 'Unnamed: 2_level_3', 'Unnamed: 2_level_4'
            ]
        """
        columns = [
            i
            for i in df.columns
            if title == i[0]
        ]
        if len(columns) != 1:
            breakpoint()
            raise Exception(f"找不到{title}")
        return columns[0]

    def get_series(self, df, columns):
        return df.loc[:, columns]
        if len(columns) == 1:
            return list(df[columns[0]])
        return self.get_series(df[columns[0]], columns[1:])

    def get_position(self, _type):
        ws = self.sheet
        result = {"column": "I"}
        e_index = column_index_from_string('E') - 1
        _type_index = column_index_from_string('B') - 1
        column_index = column_index_from_string('G') - 1
        for index, row in enumerate(ws.rows, 1):
            e_value = row[e_index].value
            if e_value \
                    and "广汽丰田" == e_value \
                    and row[_type_index].value == _type:
                if row[column_index].value in [
                        "收货数", "发货数MP",
                        "不良品返厂",
                ]:
                    result[row[column_index].value] = index
        logger.info(f"{_type}的数据粘贴位置： {result}")
        return result


paste_convert = Convert.paste_convert
